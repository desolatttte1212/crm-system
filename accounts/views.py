from django.core.checks import messages
from django.http import JsonResponse
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate, login
from django import forms
from .models import Request
from .models import Notification
from .password_validators import validate_strong_password
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.models import User, Group
from django.http import HttpResponse
from .models import Profile
from django.core.exceptions import ValidationError as DjangoValidationError
import logging
from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import get_object_or_404
from django.db.models import Sum
from django.db.models import Q

# Форма регистрации
class RegisterForm(forms.Form):
    username = forms.CharField(max_length=150, label="Имя пользователя")
    email = forms.EmailField(label="Email")
    password = forms.CharField(widget=forms.PasswordInput, label="Пароль")
    password2 = forms.CharField(widget=forms.PasswordInput, label="Повторите пароль")
    policy_accepted = forms.BooleanField(required=True, label="Я принимаю политику конфиденциальности")

    def clean(self):
        cleaned_data = super().clean()
        password = cleaned_data.get("password")
        password2 = cleaned_data.get("password2")
        if password and password2 and password != password2:
            raise forms.ValidationError("Пароли не совпадают")
        return cleaned_data


# Форма входа
class LoginForm(forms.Form):
    username = forms.CharField(label="Имя пользователя")
    hashed_password = forms.CharField(widget=forms.HiddenInput())


# Представления
logger = logging.getLogger(__name__)

def validate_strong_password(password):
    errors = []
    if len(password) < 8:
        errors.append("Пароль должен быть не менее 8 символов.")
    if not any(c.isupper() for c in password):
        errors.append("Пароль должен содержать хотя бы одну заглавную букву.")
    if not any(c.isdigit() for c in password):
        errors.append("Пароль должен содержать хотя бы одну цифру.")
    if not any(c in '!@#$%^&*()_+-=' for c in password):
        errors.append("Пароль должен содержать хотя бы один специальный символ.")
    if errors:
        raise DjangoValidationError(errors)


def secure_register_view(request):
    if request.method == 'POST':
        print("\n" + "=" * 50)
        print("🔹 secure_register_view: POST получен")
        print("POST-ключи:", list(request.POST.keys()))

        form = RegisterForm(request.POST)

        if form.is_valid():
            print("✅ Форма валидна. Данные:", form.cleaned_data)

            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Проверка на существование
            if User.objects.filter(username=username).exists():
                print(f"❌ Пользователь с логином '{username}' уже существует")
                form.add_error('username', 'Пользователь с таким логином уже существует.')
            elif User.objects.filter(email=email).exists():
                print(f"❌ Пользователь с email '{email}' уже существует")
                form.add_error('email', 'Пользователь с таким email уже существует.')
            else:
                try:
                    # Проверка сложности пароля
                    print("🔹 Проверка сложности пароля...")
                    validate_strong_password(password)
                    print("✅ Пароль прошёл проверку")

                    # Создаём пользователя
                    print("🔹 Создаём пользователя...")
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password
                    )
                    print(f"✅ Пользователь создан: ID={user.id}, username='{user.username}'")
                    # Создаём профиль
                    try:
                        print("🔹 Создаём профиль для пользователя...")
                        Profile.objects.create(user=user)
                        print("✅ Профиль успешно создан")
                    except Exception as e:
                        print(f"❌ Ошибка при создании профиля: {str(e)}")
                        form.add_error(None, "Не удалось создать профиль пользователя.")
                        return render(request, 'accounts/register_secure.html', {'form': form})

                    # Автоматический вход
                    print("🔹 Выполняем вход пользователя...")
                    login(request, user)
                    print("✅ Пользователь вошёл в систему")

                    # Успешная регистрация
                    messages.success(request, "Вы успешно зарегистрировались!")
                    print("🚀 РЕДИРЕКТ: Перенаправляем на 'my_requests'")
                    return redirect('home')

                except DjangoValidationError as e:
                    print(f"❌ Ошибка валидации пароля: {e.messages}")
                    for error in e.messages:
                        form.add_error('password', error)
                except Exception as e:
                    print(f"❌ Критическая ошибка при регистрации: {type(e).__name__}: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    logger.error(f"Ошибка при регистрации: {str(e)}")
                    form.add_error(None, "Произошла ошибка. Попробуйте позже.")
        else:
            print("❌ Форма НЕ валидна")
            print("Ошибки формы:", dict(form.errors))
    else:
        form = RegisterForm()

    return render(request, 'accounts/register_secure.html', {'form': form})

def login_view(request):
    error = None

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)

            # 🔍 Проверяем группу и перенаправляем
            if user.groups.filter(name='System Admin').exists():
                return redirect('system_admin_panel')
            if user.groups.filter(name='Lead Managers').exists():
                return redirect('lead_manager_page')
            elif user.groups.filter(name='Managers').exists():
                return redirect('manager_requests')
            elif user.groups.filter(name='CTO Department').exists():
                return redirect('cto_department')
            elif user.groups.filter(name='Economist').exists():
                return redirect('economist_dashboard')
            elif user.groups.filter(name="CEO").exists():
                return redirect("ceo_dashboard")
            elif user.groups.filter(name="Accountant").exists():
                return redirect("accountant_dashboard")
            elif user.groups.filter(name='Production').exists():
                return redirect('production_dashboard')
            elif user.groups.filter(name='Warehouse Manager').exists():
                return redirect('warehouse_manager')
            else:
                # Клиент
                return redirect('home')
        else:
            error = "Неверное имя пользователя или пароль"

    return render(request, 'accounts/login.html', {
        'error': error
    })

def logout_view(request):
    auth_logout(request)
    return redirect('login')

@login_required
def create_request_view(request):
    # Очищаем любые старые сообщения при загрузке страницы
    storage = messages.get_messages(request)
    for message in storage:
        pass  # Просто читаем сообщения чтобы очистить их

    if not hasattr(request.user, 'profile') or not request.user.profile.is_profile_complete:
        messages.warning(request, "Заполните профиль.")
        return redirect('complete_profile')

    error = None
    if request.method == 'POST':
        product_id = request.POST.get('product_id', '').strip()
        quantity_str = request.POST.get('quantity', '').strip()
        delivery_type = request.POST.get('delivery_type')

        # Валидация
        if not product_id:
            error = "Выберите товар."
        elif not quantity_str.isdigit() or int(quantity_str) <= 0:
            error = "Введите корректное количество."
        elif delivery_type not in ['pickup', 'delivery']:
            error = "Укажите способ доставки."

        if not error:
            try:
                # Получаем товар из базы
                product = get_object_or_404(Product, id=product_id, is_available=True)
                quantity = int(quantity_str)

                # Проверяем наличие товара
                if product.quantity < quantity:
                    error = f"Недостаточно товара на складе. Доступно: {product.quantity}"
                else:
                    # Находим любого пользователя из группы "Lead Managers"
                    lead_manager = User.objects.filter(groups__name='Lead Managers').first()
                    if not lead_manager:
                        error = "Нет доступных главных менеджеров."
                    else:
                        # Создаем заявку
                        request_obj = Request.objects.create(
                            user=request.user,
                            product_name=product.name,
                            product=product,  # Связываем с реальным товаром
                            quantity=quantity,
                            delivery_type=delivery_type,
                            status='sent',
                            assigned_manager=lead_manager
                        )

                        # Резервируем товар
                        product.quantity -= quantity
                        if product.quantity == 0:
                            product.is_available = False
                        product.save()

                        # Логируем списание
                        InventoryLog.objects.create(
                            product=product,
                            movement_type='outgoing',
                            quantity=quantity,
                            description=f"Зарезервировано для заявки №{request_obj.id}",
                            user=request.user
                        )

                        # Уведомление
                        Notification.objects.create(
                            user=lead_manager,
                            message=f"Новая заявка №{request_obj.id} на '{product.name}' ожидает распределения"
                        )

                        return redirect('my_requests')  # Просто редирект без сообщения

            except Product.DoesNotExist:
                error = "Товар недоступен."
            except Exception as e:
                print(f"Ошибка: {e}")
                error = "Ошибка при создании заявки."

    # Получаем доступные товары из базы
    available_products = Product.objects.filter(is_available=True, quantity__gt=0)

    return render(request, 'accounts/create_request.html', {
        'available_products': available_products,
        'error': error
    })


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Economist').exists())
def economist_cancel(request, request_id):
    if request.method == 'POST':
        req = get_object_or_404(Request, id=request_id)
        old_status = req.status
        req.status = 'cancelled'
        req.save()

        # Возвращаем товар на склад
        return_product_to_stock(req, old_status)

        messages.success(request, f"Заявка №{req.id} отменена, товар возвращен на склад")

    return redirect('economist_dashboard')

def return_product_to_stock(request_obj, old_status=None):
    """Возвращает товар на склад при отмене заявки"""
    print(f"🔍 DEBUG: Попытка возврата товара для заявки {request_obj.id}")
    print(f"🔍 DEBUG: Статус заявки: {request_obj.status}")
    print(f"🔍 DEBUG: Старый статус: {old_status}")
    print(f"🔍 DEBUG: Есть ли товар: {request_obj.product is not None}")

    if request_obj.product:
        print(f"🔍 DEBUG: Товар: {request_obj.product.name}")
        print(f"🔍 DEBUG: Количество в заявке: {request_obj.quantity}")
        print(f"🔍 DEBUG: Текущее количество на складе: {request_obj.product.quantity}")

        # Разрешаем возврат для статусов, которые были ДО отмены
        allowed_old_statuses = [
            'sent', 'processing', 'awaiting_client', 'awaiting_cost',
            'awaiting_tkp', 'awaiting_approval', 'awaiting_documents'
        ]

        # Используем старый статус, если передан, иначе текущий
        status_to_check = old_status if old_status else request_obj.status

        if status_to_check in allowed_old_statuses:
            try:
                product = request_obj.product
                old_quantity = product.quantity
                product.quantity += request_obj.quantity

                if product.quantity > 0:
                    product.is_available = True

                product.save()
                print(f"✅ SUCCESS: Товар возвращен! Было: {old_quantity}, Стало: {product.quantity}")

                # Словарь для перевода статусов на русский
                status_translations = {
                    'sent': 'Отправлена',
                    'processing': 'В обработке',
                    'awaiting_client': 'На согласовании у клиента',
                    'awaiting_cost': 'Ожидание расчёта стоимости',
                    'awaiting_tkp': 'Ожидание ТКП',
                    'awaiting_approval': 'На согласовании',
                    'awaiting_documents': 'На оформлении документов'
                }

                # Получаем русское название статуса
                status_russian = status_translations.get(status_to_check, status_to_check)

                # Создаем запись в журнале с русским статусом
                log_entry = InventoryLog.objects.create(
                    product=product,
                    movement_type='incoming',
                    quantity=request_obj.quantity,
                    description=f"Возврат из отмененной заявки №{request_obj.id} (был статус: {status_russian})",
                    user=request_obj.assigned_manager or request_obj.user
                )
                print(f"✅ SUCCESS: Запись в журнале создана: {log_entry}")

            except Exception as e:
                print(f"❌ ERROR: Ошибка при возврате товара: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"⚠️ WARNING: Статус {status_to_check} не позволяет вернуть товар")
    else:
        print("⚠️ WARNING: Нет связанного товара для возврата")

@login_required
def home_view(request):
    # Уведомления
    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    # Статистика для клиента
    active_statuses = [
        'sent', 'processing', 'awaiting_client', 'awaiting_cost',
        'awaiting_documents', 'signed_by_client', 'shipping_docs_ready', 'client_signed_shipping'
    ]
    completed_statuses = ['completed']
    cancelled_statuses = ['cancelled']

    my_requests = Request.objects.filter(user=request.user)

    total_active = my_requests.filter(status__in=active_statuses).count()
    total_completed = my_requests.filter(status__in=completed_statuses).count()
    total_cancelled = my_requests.filter(status__in=cancelled_statuses).count()

    # Последняя заявка
    last_request = my_requests.order_by('-created_at').first()

    # Общая сумма завершённых заказов
    total_cost = my_requests.filter(status='completed').aggregate(
        total=Sum('cost_estimate')
    )['total'] or 0

    # Передаём статистику
    stats = {
        'total_active': total_active,
        'total_completed': total_completed,
        'total_cancelled': total_cancelled,
        'last_request': last_request,
        'total_cost': total_cost,
    }

    context = {
        'notifications': notifications,
        'unread_notifications': unread_notifications,
        'stats': stats,
    }

    # Редирект по ролям
    user = request.user

    if user.groups.filter(name='Managers').exists():
        return redirect('manager_requests')
    elif user.groups.filter(name='Lead Managers').exists():
        return redirect('lead_manager_page')
    elif user.groups.filter(name='Accountant').exists():
        return redirect('accountant_dashboard')
    elif user.groups.filter(name='Production').exists():
        return redirect('production_dashboard')
    elif user.groups.filter(name='CEO').exists():
        return redirect('ceo_dashboard')
    elif user.groups.filter(name='CTO Department').exists():
        return redirect('cto_department')
    elif user.groups.filter(name='Economist').exists():
        return redirect('economist_dashboard')
    elif user.groups.filter(name='System Admin').exists():
        return redirect('system_admin_panel')
    elif user.groups.filter(name='Warehouse Manager').exists():
        return redirect('warehouse_manager')

    # Для клиента — показываем home
    return render(request, 'accounts/home.html', context)

def profile_view(request):
    user = request.user
    profile, created = Profile.objects.get_or_create(user=user)
    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    if request.method == 'POST':
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        company = request.POST.get('company', '').strip()

        profile.phone = phone
        profile.address = address
        profile.company = company
        profile.save()

        return redirect('profile')

    return render(request, 'accounts/profile.html', {
        'user': user,
        'profile': profile,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })


def my_requests_view(request):
    requests = Request.objects.filter(user=request.user).order_by('-created_at')
    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()
    return render(request, 'accounts/my_requests.html', {
        'requests': requests,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })

def add_comment_view(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        comment = request.POST.get('comment')
        req.manager_comment = comment
        req.save()

        # Уведомление клиенту
        Notification.objects.create(
            user=req.user,
            message=f"Новый комментарий по заявке '{req.product_name}': {comment}"
        )
    return redirect('manager_requests')

@login_required
def mark_notifications_as_read(request):
    if request.method == 'POST':
        # Помечаем все непрочитанные уведомления как прочитанные
        unread = request.user.notifications.filter(is_read=False)
        count = unread.count()
        unread.update(is_read=True)
        return JsonResponse({'status': 'success', 'cleared': count})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def delete_notification(request, notification_id):
    note = get_object_or_404(Notification, id=notification_id, user=request.user)
    note.delete()
    return JsonResponse({'status': 'success'})


from django.contrib.auth.decorators import login_required, user_passes_test


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def manager_requests_view(request):
    # Начинаем с активных заявок (не в архиве)
    requests = Request.objects.filter(
        assigned_manager=request.user
    ).exclude(
        status__in=['completed', 'cancelled']
    ).order_by('-created_at')

    # Сохраняем параметры для сортировки и пагинации
    filters = request.GET.copy()
    if 'page' in filters:
        del filters['page']
    request.session['filters'] = filters

    # Поиск по товару
    search = request.GET.get('search')
    if search:
        requests = requests.filter(product_name__icontains=search)

    # Поиск по клиенту
    client = request.GET.get('client')
    if client:
        requests = requests.filter(user__username__icontains=client) | \
                   requests.filter(user__profile__company_name__icontains=client)

    # Фильтр по статусу
    status_filter = request.GET.get('status')
    if status_filter:
        requests = requests.filter(status=status_filter)

    # Фильтр по дате
    date_from = request.GET.get('date_from', '')  # ← Пустая строка, а не None
    date_to = request.GET.get('date_to', '')
    if date_from:
        requests = requests.filter(created_at__date__gte=date_from)
    if date_to:
        requests = requests.filter(created_at__date__lte=date_to)

    # Сортировка
    sort = request.GET.get('sort')
    if sort == 'date_asc':
        requests = requests.order_by('created_at')
    elif sort == 'date_desc':
        requests = requests.order_by('-created_at')

    # Для передачи в шаблон
    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/manager_requests.html', {
        'requests': requests,
        'search': search or '',
        'client': client or '',
        'status_filter': status_filter or '',
        'date_from': date_from or '',
        'date_to': date_to or '',
        'sort': sort or '',
        'filters': filters.urlencode(),
        'notifications': notifications,
        'unread_notifications': unread_notifications,
        'STATUS_CHOICES': Request.STATUS_CHOICES,
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Lead Managers').exists())
def assign_request_to_manager(request, request_id):
    req = get_object_or_404(Request, id=request_id)

    if request.method == 'POST':
        manager_id = request.POST.get('manager_id')

        old_manager = req.assigned_manager

        # Если выбрано "Снять с менеджера"
        if manager_id == 'none':
            req.assigned_manager = None
            req.save()
            if old_manager:
                Notification.objects.create(
                    user=old_manager,
                    message=f"Вы откреплены от заявки №{req.id}"
                )
        # Если выбран конкретный менеджер
        elif manager_id:
            manager = get_object_or_404(User, id=manager_id, groups__name='Managers')
            req.assigned_manager = manager
            req.save()

            # Уведомление новому менеджеру
            Notification.objects.create(
                user=manager,
                message=f"Вам назначена заявка №{req.id} на '{req.product_name}'"
            )

            # Уведомление старому менеджеру (если был)
            if old_manager and old_manager != manager:
                Notification.objects.create(
                    user=old_manager,
                    message=f"Вы откреплены от заявки №{req.id}"
                )

    return redirect('lead_manager_page')

from django.db.models import Q

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Lead Managers').exists())
def lead_manager_dashboard(request):
    # Показываем ВСЕ заявки
    requests = Request.objects.all().select_related('user', 'assigned_manager').order_by('-created_at')

    # Фильтры
    search = request.GET.get('search')
    client = request.GET.get('client')
    status_filter = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if search:
        requests = requests.filter(product_name__icontains=search)
    if client:
        requests = requests.filter(user__username__icontains=client)
    if status_filter:
        requests = requests.filter(status=status_filter)
    if date_from:
        requests = requests.filter(created_at__date__gte=date_from)
    if date_to:
        requests = requests.filter(created_at__date__lte=date_to)

    # Сортировка
    sort = request.GET.get('sort')
    if sort == 'date_asc':
        requests = requests.order_by('created_at')
    elif sort == 'date_desc':
        requests = requests.order_by('-created_at')

    # Уведомления
    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    # Все менеджеры для выпадающего списка
    managers = User.objects.filter(groups__name='Managers')

    return render(request, 'accounts/lead_manager_page.html', {
        'requests': requests,
        'managers': managers,
        'search': search or '',
        'client': client or '',
        'status_filter': status_filter or '',
        'date_from': date_from or '',
        'date_to': date_to or '',
        'sort': sort or '',
        'filters': request.GET.urlencode(),
        'notifications': notifications,
        'unread_notifications': unread_notifications,
        'STATUS_CHOICES': Request.STATUS_CHOICES,
    })
from django.contrib.auth.decorators import login_required, user_passes_test

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Lead Managers').exists())
def lead_manager_clients(request):
    # Все пользователи, которые НЕ в группах Managers или Lead Managers
    clients = User.objects.filter(
        groups__isnull=True
    ).prefetch_related('profile').order_by('-date_joined')

    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/lead_manager_clients.html', {
        'clients': clients,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def manager_update_request_status(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        old_status = req.status
        status = request.POST.get('status')

        if status in dict(Request.STATUS_CHOICES):
            req.status = status
            req.save()

            # Возвращаем товар при отмене
            if status == 'cancelled' and old_status != 'cancelled':
                return_product_to_stock(req, old_status)  # ПЕРЕДАЕМ СТАРЫЙ СТАТУС

            # Уведомление клиенту
            Notification.objects.create(
                user=req.user,
                message=f"Статус вашей заявки '{req.product_name}' изменён на '{req.get_status_display()}'"
            )

    return redirect('manager_requests')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='CEO').exists())
def cto_update_request_status(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        status = request.POST.get('status')
        # Разрешённые статусы для КТО
        allowed_statuses = ['awaiting_documents', 'documents_ready']
        if status in allowed_statuses:
            req.status = status
            req.save()

            Notification.objects.create(
                user=req.assigned_manager,
                message=f"Директор обновил статус заявки '{req.product_name}'"
            )

    return redirect('ceo_dashboard')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def add_comment_view(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        comment = request.POST.get('comment')
        req.manager_comment = comment
        req.save()

        Notification.objects.create(
            user=req.user,
            message=f"Новый комментарий по заявке '{req.product_name}': {comment}"
        )
    return redirect('manager_requests')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists(), login_url='home')
def manager_clients(request):
    # Получаем клиентов, у которых есть заявки, назначенные текущему менеджеру
    clients = User.objects.filter(
        request__assigned_manager=request.user
    ).distinct().order_by('username')

    # Добавляем количество заявок для каждого клиента
    clients_with_stats = []
    for client in clients:
        active_requests = Request.objects.filter(
            user=client,
            assigned_manager=request.user
        ).exclude(status__in=['completed', 'cancelled']).count()

        total_requests = Request.objects.filter(
            user=client,
            assigned_manager=request.user
        ).count()

        clients_with_stats.append({
            'client': client,
            'active_requests': active_requests,
            'total_requests': total_requests,
        })

    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/manager_clients.html', {
        'clients_with_stats': clients_with_stats,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Lead Managers').exists())
def unassign_request(request, request_id):
    """
    Снимает назначение менеджера с заявки
    """
    req = get_object_or_404(Request, id=request_id)
    req.assigned_manager = None
    req.save()

    # Уведомление клиенту
    Notification.objects.create(
        user=req.user,
        message=f"Ваша заявка '{req.product_name}' временно снята с обработки"
    )

    return redirect('lead_manager_page')

from django.contrib.auth.models import Group

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Lead Managers').exists())
def transfer_request(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        new_status = request.POST.get('new_status')
        if new_status in dict(Request.STATUS_CHOICES):
            old_status = req.status
            req.status = new_status
            req.save()

            # Уведомление клиенту
            Notification.objects.create(
                user=req.user,
                message=f"Статус вашей заявки '{req.product_name}' изменён на '{req.get_status_display()}'"
            )

            # Уведомление следующего отдела
            if new_status == 'awaiting_tkp':
                # Отправляем уведомление в КТО
                cto_group = Group.objects.get(name='CTO Department')
                for user in cto_group.user_set.all():
                    Notification.objects.create(
                        user=user,
                        message=f"Новая заявка '{req.product_name}' требует вашего внимания"
                    )
            elif new_status == 'awaiting_approval':
                # Отправляем уведомление гендиректору
                ceo_group = Group.objects.get(name='CEO')
                for user in ceo_group.user_set.all():
                    Notification.objects.create(
                        user=user,
                        message=f"Заявка '{req.product_name}' требует вашего согласования"
                    )

    return redirect('lead_manager_page')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Commercial Department').exists())
def commercial_department_view(request):
    requests = Request.objects.filter(status='awaiting_tkp').order_by('-created_at')

    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/commercial_department.html', {
        'requests': requests,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='System Admin').exists(), login_url='home')
def system_admin_panel(request):
    # Все пользователи, кроме суперпользователей
    users = User.objects.filter(is_superuser=False).select_related('profile').order_by('username')
    groups = Group.objects.all().order_by('name')  # Все группы

    return render(request, 'accounts/system_admin_panel.html', {
        'users': users,
        'groups': groups,
    })


@login_required
@user_passes_test(lambda u: u.groups.filter(name='System Admin').exists())
def set_user_role(request, user_id):
    user = get_object_or_404(User, id=user_id, is_superuser=False)
    if request.method == 'POST':
        user.groups.clear()
        group_id = request.POST.get('group')
        if group_id:
            group = get_object_or_404(Group, id=group_id)
            user.groups.add(group)

    return redirect('system_admin_panel')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='CTO Department').exists(), login_url='home')
def cto_department_view(request):
    requests = Request.objects.filter(status='awaiting_approval').order_by('-created_at')

    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/cto_department.html', {
        'requests': requests,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='CTO Department').exists())
def cto_approve_for_cost(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        comment = request.POST.get('cto_comment', '').strip()
        req.cto_comment = comment
        req.status = 'awaiting_cost'  # Только после одобрения
        req.save()

        # Уведомление экономисту
        try:
            economist_group = Group.objects.get(name='Economist')
            for user in economist_group.user_set.all():
                Notification.objects.create(
                    user=user,
                    message=f"Новая заявка '{req.product_name}' требует расчёта стоимости"
                )
        except Group.DoesNotExist:
            pass

    return redirect('cto_department')


@login_required
@user_passes_test(lambda u: u.groups.filter(name='CTO Department').exists())
def cto_reject(request, request_id):
    print(f"🔍 DEBUG: Вызов cto_reject для заявки {request_id}")

    req = get_object_or_404(Request, id=request_id)
    print(f"🔍 DEBUG: Заявка найдена: {req.id}, текущий статус: {req.status}")
    print(f"🔍 DEBUG: Связанный товар: {req.product}")

    if req.product:
        print(f"🔍 DEBUG: Товар: {req.product.name}, количество: {req.quantity}")

    old_status = req.status  # Сохраняем старый статус ДО изменения
    req.status = 'cancelled'
    req.save()
    print(f"🔍 DEBUG: Статус изменен с '{old_status}' на 'cancelled'")

    # Возвращаем товар на склад, передавая старый статус
    print("🔍 DEBUG: Вызов return_product_to_stock...")
    return_product_to_stock(req, old_status)

    # Уведомление менеджеру
    Notification.objects.create(
        user=req.assigned_manager,
        message=f"Заявка '{req.product_name}' отклонена КТО"
    )

    return redirect('cto_department')


@login_required
@user_passes_test(lambda u: u.groups.filter(name='CTO Department').exists())
def cto_add_comment(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        comment = request.POST.get('cto_comment', '').strip()
        req.cto_comment = comment

        # Если комментарий добавлен, но заявка не одобрена, можно отменить
        if 'reject' in request.POST:  # Если есть кнопка отклонения
            old_status = req.status  # СОХРАНЯЕМ СТАРЫЙ СТАТУС
            req.status = 'cancelled'

            # Возвращаем товар на склад
            return_product_to_stock(req, old_status)  # ПЕРЕДАЕМ СТАРЫЙ СТАТУС

            Notification.objects.create(
                user=req.assigned_manager,
                message=f"Заявка '{req.product_name}' отклонена КТО с комментарием"
            )
        else:
            req.status = 'awaiting_cost'
            req.save()
            try:
                economist_group = Group.objects.get(name='Economist')
                for user in economist_group.user_set.all():
                    Notification.objects.create(
                        user=user,
                        message=f"Новая заявка '{req.product_name}' требует расчёта стоимости"
                    )
            except Group.DoesNotExist:
                pass

    return redirect('cto_department')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Economist').exists())
def economist_view(request):
    # Принудительно очищаем ВСЕ сообщения
    storage = messages.get_messages(request)
    for message in storage:
        pass  # Очищаем все сообщения

    # очищаем storage полностью
    storage.used = True

    requests = Request.objects.filter(status='awaiting_cost').order_by('-created_at')
    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/economist_dashboard.html', {
        'requests': requests,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Economist').exists())
def economist_update_cost(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        cost = request.POST.get('cost_estimate', '').strip()
        delivery = request.POST.get('delivery_estimate', '').strip()

        # Валидация - проверяем, что заполнены оба поля
        if not cost or not delivery:
            messages.error(request, "Заполните все поля: стоимость и сроки поставки")
            return redirect('economist_dashboard')

        try:
            # Преобразуем стоимость в число
            req.cost_estimate = float(cost)
            req.delivery_estimate = delivery

            # Переводим в статус "На согласовании у клиента"
            req.status = 'awaiting_client'
            req.save()

            # Уведомление менеджеру и клиенту
            Notification.objects.create(
                user=req.assigned_manager,
                message=f"Заявка '{req.product_name}' ожидает согласования с клиентом"
            )
            Notification.objects.create(
                user=req.user,
                message=f"Ваша заявка '{req.product_name}' ожидает вашего согласования"
            )

            messages.success(request, "Расчёт стоимости отправлен клиенту")

        except (ValueError, TypeError):
            messages.error(request, "Некорректное значение стоимости")

    return redirect('economist_dashboard')

from django.utils import timezone

@login_required
def client_approve(request, request_id):
    req = get_object_or_404(Request, id=request_id, user=request.user)
    if req.status == 'awaiting_client':
        req.client_approved = True
        req.client_approval_date = timezone.now()
        req.status = 'awaiting_documents'
        req.save()

        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Клиент одобрил заявку '{req.product_name}'. Начните оформление документов"
        )

    return redirect('my_requests')


@login_required
def client_reject(request, request_id):
    req = get_object_or_404(Request, id=request_id, user=request.user)
    if req.status == 'awaiting_client':
        req.client_approved = False
        req.client_approval_date = timezone.now()
        req.status = 'cancelled'
        req.client_response_received = True
        req.save()

        # Возвращаем товар на склад
        return_product_to_stock(req, 'awaiting_client')  # СТАРЫЙ СТАТУС!

        # Уведомление менеджеру
        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Клиент отклонил заявку '{req.product_name}'"
        )

    return redirect('my_requests')


logger = logging.getLogger(__name__)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def send_to_ceo_and_accountant(request, request_id):
    logger.info(f"Вызван send_to_ceo_and_accountant для request_id={request_id}")

    req = get_object_or_404(Request, id=request_id, assigned_manager=request.user)

    logger.info(f"Заявка найдена: {req.product_name}")

    try:
        ceo_group = Group.objects.get(name='CEO')
        logger.info(f"Группа CEO найдена, пользователей: {ceo_group.user_set.count()}")
        for user in ceo_group.user_set.all():
            Notification.objects.create(
                user=user,
                message=f"Подготовьте договор по заявке '{req.product_name}'"
            )
            logger.info(f"Уведомление отправлено: {user.username}")
    except Group.DoesNotExist:
        logger.warning("Группа CEO не найдена")
        pass

    try:
        acc_group = Group.objects.get(name='Accountant')
        logger.info(f"Группа Accountant найдена, пользователей: {acc_group.user_set.count()}")
        for user in acc_group.user_set.all():
            Notification.objects.create(
                user=user,
                message=f"Подготовьте счёт по заявке '{req.product_name}'"
            )
            logger.info(f"Уведомление отправлено: {user.username}")
    except Group.DoesNotExist:
        logger.warning("Группа Accountant не найдена")
        pass

    return redirect('manager_requests')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='CEO').exists())
def ceo_dashboard(request):
    # Показываем заявки, где нужно подписать договор
    requests = Request.objects.filter(status='awaiting_documents').order_by('-created_at')

    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/ceo_dashboard.html', {
        'requests': requests,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })

# Генеральный директор
@login_required
@user_passes_test(lambda u: u.groups.filter(name='CEO').exists())
def ceo_sign_contract(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST' and request.FILES.get('signed_contract_file'):
        req.signed_contract_file = request.FILES['signed_contract_file']
        req.save()

        # Уведомление менеджеру
        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Подписанный договор по заявке '{req.product_name}' готов"
        )

    return redirect('ceo_dashboard')


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Accountant').exists())
def accountant_create_invoice(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST' and request.FILES.get('invoice_file'):
        req.invoice_file = request.FILES['invoice_file']
        req.save()

        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Счёт по заявке '{req.product_name}' готов"
        )

    return redirect('accountant_dashboard')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Accountant').exists())
def accountant_dashboard(request):
    # Счёта
    invoice_requests = Request.objects.filter(status='awaiting_documents').order_by('-created_at')
    # Отгрузка
    shipping_requests = Request.objects.filter(status='ready_for_delivery').order_by('-created_at')
    # Оплата
    payment_requests = Request.objects.filter(status='awaiting_payment').order_by('-created_at')

    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/accountant_dashboard.html', {
        'invoice_requests': invoice_requests,
        'shipping_requests': shipping_requests,
        'payment_requests': payment_requests,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def send_documents_to_client(request, request_id):
    req = get_object_or_404(Request, id=request_id, assigned_manager=request.user)

    # Проверяем, что файлы прикреплены
    if req.signed_contract_file and req.invoice_file:
        req.status = 'documents_ready'  # Критически важно
        req.save()

        Notification.objects.create(
            user=req.user,
            message=f"Документы по заявке '{req.product_name}' готовы к подписанию"
        )

    return redirect('manager_requests')

@login_required
def client_sign_documents(request, request_id):
    req = get_object_or_404(Request, id=request_id, user=request.user)
    if request.method == 'POST':
        # Проверяем, что файлы загружены
        if 'client_signed_contract' in request.FILES:
            req.client_signed_contract = request.FILES['client_signed_contract']
        if 'client_signed_invoice' in request.FILES:
            req.client_signed_invoice = request.FILES['client_signed_invoice']

        # Меняем статус
        req.status = 'signed_by_client'
        req.save()  # Обязательно сохраняем

        # Уведомление менеджеру
        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Клиент подписал документы по заявке '{req.product_name}'"
        )

    return redirect('my_requests')


# accounts/views.py
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def send_to_production(request, request_id):
    req = get_object_or_404(Request, id=request_id, assigned_manager=request.user)

    # Проверяем, что документы подписаны
    if req.status == 'signed_by_client' and req.client_signed_contract and req.client_signed_invoice:
        req.status = 'in_production'
        req.save()

        Notification.objects.create(
            user=req.user,
            message=f"Заявка '{req.product_name}' передана в производство"
        )

    return redirect('manager_requests')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Production').exists(), login_url='home')
def production_dashboard(request):
    # Показываем заявки, которые в производстве
    requests = Request.objects.filter(status='in_production').order_by('-created_at')

    notifications = request.user.notifications.all().order_by('-created_at')[:10]
    unread_notifications = request.user.notifications.filter(is_read=False).count()

    return render(request, 'accounts/production_dashboard.html', {
        'requests': requests,
        'notifications': notifications,
        'unread_notifications': unread_notifications
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Production').exists())
def production_update_request_status(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST':
        status = request.POST.get('status')

        # Разрешаем только статусы для производства
        allowed_statuses = ['in_production', 'ready_for_delivery']
        if status in allowed_statuses:
            req.status = status
            req.save()

            Notification.objects.create(
                user=req.assigned_manager,
                message=f"Заявка '{req.product_name}' обновлена: {req.get_status_display()}"
            )

    return redirect('production_dashboard')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Production').exists())
def send_to_head_accountant(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if req.status == 'ready_for_delivery':
        req.status = 'awaiting_shipping_docs'
        req.save()

        # Уведомление главному бухгалтеру
        try:
            head_acc_group = Group.objects.get(name='Head Accountant')
            for user in head_acc_group.user_set.all():
                Notification.objects.create(
                    user=user,
                    message=f"Оформите отгрузочные документы по заявке '{req.product_name}'"
                )
        except Group.DoesNotExist:
            pass

    return redirect('production_dashboard')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Accountant').exists())
def accountant_add_shipping_docs(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if request.method == 'POST' and request.FILES.get('shipping_docs'):
        req.shipping_docs = request.FILES['shipping_docs']
        req.status = 'shipping_docs_ready'
        req.save()

        # Уведомление менеджеру
        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Отгрузочные документы по заявке '{req.product_name}' готовы к отправке клиенту"
        )

    return redirect('accountant_dashboard')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def send_shipping_docs_to_client(request, request_id):
    req = get_object_or_404(Request, id=request_id, assigned_manager=request.user)
    if req.status == 'shipping_docs_ready' and req.shipping_docs:
        req.status = 'shipping_docs_to_client'
        req.save()

        # Уведомление клиенту
        Notification.objects.create(
            user=req.user,
            message=f"Вам доступны отгрузочные документы по заявке '{req.product_name}'"
        )

    return redirect('manager_requests')

@login_required
def client_sign_shipping_docs(request, request_id):
    req = get_object_or_404(Request, id=request_id, user=request.user)
    if request.method == 'POST' and request.FILES.get('client_signed_shipping_docs'):
        req.client_signed_shipping_docs = request.FILES['client_signed_shipping_docs']
        req.status = 'client_signed_shipping'
        req.save()

        # Уведомление менеджеру
        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Клиент подписал отгрузочные документы по заявке '{req.product_name}'"
        )

    return redirect('my_requests')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def mark_as_awaiting_payment(request, request_id):
    req = get_object_or_404(Request, id=request_id, assigned_manager=request.user)
    if req.status == 'client_signed_shipping' and req.client_signed_shipping_docs:
        req.status = 'awaiting_payment'
        req.save()

        # Уведомление клиенту
        Notification.objects.create(
            user=req.user,
            message=f"Ваша заявка '{req.product_name}' ожидает оплаты. Пожалуйста, произведите оплату по счёту."
        )

    return redirect('manager_requests')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Accountant').exists())
def confirm_payment(request, request_id):
    req = get_object_or_404(Request, id=request_id)
    if req.status == 'awaiting_payment':
        req.status = 'completed'
        req.save()

        # Уведомление менеджеру и клиенту
        Notification.objects.create(
            user=req.assigned_manager,
            message=f"Оплата по заявке '{req.product_name}' подтверждена. Заявка завершена."
        )
        Notification.objects.create(
            user=req.user,
            message=f"Ваша заявка '{req.product_name}' завершена. Спасибо за сотрудничество!"
        )

    return redirect('accountant_dashboard')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def manager_archive(request):
    # Только завершённые и отменённые заявки, У КОТОРЫХ ЕСТЬ ПОДПИСАННЫЕ ДОКУМЕНТЫ
    requests = Request.objects.filter(
        assigned_manager=request.user,
        status__in=['completed', 'cancelled']
    ).filter(
        # Условие: хотя бы один документ от клиента загружен
        Q(client_signed_contract__isnull=False) |
        Q(client_signed_invoice__isnull=False) |
        Q(client_signed_shipping_docs__isnull=False)
    ).select_related('user__profile').order_by('-created_at')

    return render(request, 'accounts/manager_archive.html', {
        'requests': requests,
        'notifications': request.user.notifications.all()[:10],
        'unread_notifications': request.user.notifications.filter(is_read=False).count()
    })

@login_required
def complete_profile(request):
    # Гарантируем, что профиль существует
    profile, created = request.user.profile, None
    if not profile:
        profile = Profile.objects.create(user=request.user)

    if request.method == 'POST':
        client_type = request.POST.get('client_type', 'company')

        if client_type == 'company':
            profile.company_name = request.POST.get('company_name', '').strip()
            profile.full_name = ''
        else:
            profile.full_name = request.POST.get('full_name', '').strip()
            profile.company_name = ''
            profile.inn = ''

        profile.inn = request.POST.get('inn', '').strip()
        profile.phone = request.POST.get('phone', '').strip()
        profile.address = request.POST.get('address', '').strip()

        if profile.phone:
            profile.is_profile_complete = True
            profile.save()
            messages.success(request, "Профиль успешно обновлён!")
            return redirect('profile')  # Возвращаемся в личный кабинет
        else:
            messages.error(request, "Телефон обязателен.")

    return render(request, 'accounts/complete_profile.html', {'profile': profile})

@login_required
def client_response(request, request_id):
    req = get_object_or_404(Request, id=request_id, user=request.user)

    if req.client_response_received:
        messages.warning(request, "Вы уже ответили.")
        return redirect('my_requests')

    if request.method == 'POST':
        response = request.POST.get('response')
        reason = request.POST.get('reason', '').strip()

        if response == 'accepted':
            req.client_approved = True
            req.client_approval_date = timezone.now()
            req.status = 'awaiting_documents'
            messages.success(request, "Вы согласились.")
            Notification.objects.create(
                user=req.assigned_manager,
                message=f"Клиент согласился с расчёта по заявке '{req.product_name}'"
            )
        elif response == 'rejected':
            if not reason:
                messages.error(request, "Укажите причину.")
                return redirect('my_requests')

            req.client_approved = False
            req.client_rejection_reason = reason
            req.status = 'cancelled'
            req.client_response_received = True
            messages.success(request, "Вы отклонили расчёт.")

            # Возвращаем товар на склад
            return_product_to_stock(req, 'awaiting_client')  # СТАРЫЙ СТАТУС!

            Notification.objects.create(
                user=req.assigned_manager,
                message=f"Клиент отказался от заявки '{req.product_name}' — {reason[:50]}..."
            )
        else:
            return redirect('my_requests')

        req.client_response_received = True
        req.save()

    return redirect('my_requests')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Managers').exists())
def export_requests_excel(request):
    requests = Request.objects.filter(assigned_manager=request.user).select_related('user__profile')

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if date_from:
        requests = requests.filter(created_at__date__gte=date_from)
    if date_to:
        requests = requests.filter(created_at__date__lte=date_to)

    # Создаём Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Заявки"

    headers = [
        'ID', 'Товар', 'Кол-во', 'Стоимость', 'Статус', 'Дата создания',
        'Клиент', 'Телефон', 'Email', 'ИНН', 'Адрес', 'Доставка', 'Причина отмены'
    ]

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = border

    # Данные
    for req in requests:
        profile = req.user.profile if hasattr(req.user, 'profile') else None
        client_name = (
            profile.company_name or profile.full_name or req.user.username
            if profile else req.user.username
        )
        row = [
            req.id,
            req.product_name,
            req.quantity,
            f"{req.cost_estimate} ₽" if req.cost_estimate else "",
            req.get_status_display(),
            req.created_at.strftime("%d.%m.%Y %H:%M"),
            client_name,
            profile.phone if profile else "",
            req.user.email,
            profile.inn if profile else "",
            profile.address if profile else "",
            "Самовывоз" if req.delivery_type == 'pickup' else "Доставка",
            req.client_rejection_reason if req.client_rejection_reason else ""
        ]
        sheet.append(row)

    # Автоподбор ширины
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column].width = adjusted_width

    # Стили для данных
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # Имя файла
    today = datetime.now().strftime("%Y-%m-%d")
    if date_from and date_to:
        filename = f"заявки_{date_from}_до_{date_to}.xlsx"
    elif date_from:
        filename = f"заявки_с_{date_from}.xlsx"
    elif date_to:
        filename = f"заявки_до_{date_to}.xlsx"
    else:
        filename = f"заявки_{today}.xlsx"

    # Ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Disposition'] += f'; filename*=UTF-8\'\'{filename}'
    response['Cache-Control'] = 'no-cache'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'

    workbook.save(response)
    return response

from .models import Product

@login_required
def warehouse_view(request):
    products = Product.objects.filter(is_available=True)

    context = {
        'products': products,
        'notifications': request.user.notifications.all()[:10],
        'unread_notifications': request.user.notifications.filter(is_read=False).count()
    }
    return render(request, 'accounts/warehouse.html', context)

from .models import Product, InventoryLog

def is_warehouse_manager(user):
    return user.groups.filter(name='Warehouse Manager').exists()

@login_required
@user_passes_test(is_warehouse_manager)
def warehouse_manager_view(request):
    products = Product.objects.all()
    logs = InventoryLog.objects.select_related('product', 'user').order_by('-created_at')[:50]

    context = {
        'products': products,
        'logs': logs,
        'notifications': request.user.notifications.all()[:10],
        'unread_notifications': request.user.notifications.filter(is_read=False).count()
    }
    return render(request, 'accounts/warehouse_manager.html', context)

@login_required
@user_passes_test(is_warehouse_manager)
def add_product(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        price = request.POST.get('price')
        quantity = request.POST.get('quantity', 0)

        if not name or not price:
            messages.error(request, "Название и цена обязательны.")
            return redirect('warehouse_manager')

        try:
            price = float(price)
            quantity = int(quantity)
        except:
            messages.error(request, "Цена и количество должны быть числами.")
            return redirect('warehouse_manager')

        product = Product.objects.create(
            name=name,
            description=description,
            price=price,
            quantity=quantity,
            is_available=quantity > 0
        )

        # Запись в журнал
        InventoryLog.objects.create(
            product=product,
            movement_type='incoming',
            quantity=quantity,
            description="Добавление нового товара",
            user=request.user
        )

        messages.success(request, f"Товар '{product.name}' добавлен.")
        return redirect('warehouse_manager')

    return redirect('warehouse_manager')

@login_required
@user_passes_test(is_warehouse_manager)
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        old_quantity = product.quantity
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        price = request.POST.get('price')
        new_quantity = request.POST.get('quantity', 0)

        if not name or not price:
            messages.error(request, "Название и цена обязательны.")
            return redirect('warehouse_manager')

        try:
            price = float(price)
            new_quantity = int(new_quantity)
        except:
            messages.error(request, "Цена и количество должны быть числами.")
            return redirect('warehouse_manager')

        # Обновляем товар
        product.name = name
        product.description = description
        product.price = price
        product.quantity = new_quantity
        product.is_available = new_quantity > 0
        product.save()

        # Логируем изменение количества
        diff = new_quantity - old_quantity
        if diff != 0:
            movement_type = 'incoming' if diff > 0 else 'outgoing'
            InventoryLog.objects.create(
                product=product,
                movement_type=movement_type,
                quantity=abs(diff),
                description=f"Изменение количества через редактирование",
                user=request.user
            )

        messages.success(request, f"Товар '{product.name}' обновлён.")
        return redirect('warehouse_manager')

    return redirect('warehouse_manager')

@login_required
@user_passes_test(is_warehouse_manager)
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        name = product.name
        product.delete()
        messages.success(request, f"Товар '{name}' удалён.")
        return redirect('warehouse_manager')

    return redirect('warehouse_manager')
